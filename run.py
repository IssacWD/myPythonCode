# !/usr/bin/python3
# coding=utf-8
# ********************************************************
#   > OS     : OS X 10.11.3
#   > Author : JasonGUTU
#   > Mail   : hellojasongu@gmail.com
#   > Time   : 2016/4/27
# ********************************************************
import threading
import pickle
import queue

import openpyxl
import requests


def load_workbook():
    workbook = openpyxl.load_workbook(FILE_NAME)  # load workbook file
    sheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])  # get the sheet
    rows = sheet.rows
    for i in range(1, len(rows)):
        url = rows[i][6].value
        url_package = [i, url, 0]
        url_queue.put(url_package)


class RequestTest(threading.Thread):
    """ test if can be request """
    def run(self):
        while True:
            if url_queue.qsize() != 0:  # if url in queue
                try: url_package = url_queue.get()
                except: continue
                if not url_package[1].startswith('http://'):
                    url = 'http://' + url_package[1]
                else:
                    url = url_package[1]
                try:
                    request = requests.get(url)
                    url_package[2] = request.status_code
                    request.close()
                except:  # Error need to be reviewed
                    pass
                requested_url_list.append(url_package)
                print('%s,%s,%s' % (url_package[0], url_package[2], url_package[1]))
            else:
                break


if __name__ == '__main__':
    FILE_NAME = 'URL.xlsx'
    url_queue = queue.Queue()
    requested_url_list = list()
    number_of_threads = 100
    load_workbook()
    threads_pool = list()
    for i in range(number_of_threads):
        t = RequestTest()
        threads_pool.append(t)
    for thread in threads_pool:
        thread.start()
    for thread in threads_pool:
        thread.join()
    list_file = open('result.list.plk', 'wb')
    requested_url_list.sort()
    pickle.dump(requested_url_list, list_file)
